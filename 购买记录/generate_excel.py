"""
购买记录 Excel 表格生成器
包含：明细记录、月度统计、年度统计、平台分布，并带图表
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

OUTPUT_DIR = "/workspace/购买记录/归档表格"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "购买记录.xlsx")

# ==================== 样式定义 ====================
# 颜色方案
COLOR_HEADER = "2F5496"      # 深蓝（表头）
COLOR_HEADER_FONT = "FFFFFF" # 白色字
COLOR_SUBHEADER = "DEEBF7"   # 浅蓝（副标题）
COLOR_TOTAL = "FFF2CC"       # 浅黄（合计行）
COLOR_BORDER = "8EAADB"      # 边框蓝
COLOR_MONTH_ROW = "E2EFDA"   # 浅绿（月份行）
COLOR_TITLE = "1F4E79"       # 标题深蓝

# 字体
FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="微软雅黑", size=11, bold=True, color=COLOR_HEADER_FONT)
FONT_NORMAL = Font(name="微软雅黑", size=10, color="333333")
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True, color="333333")
FONT_TOTAL = Font(name="微软雅黑", size=11, bold=True, color="C00000")

# 填充
FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_SUBHEADER = PatternFill("solid", fgColor=COLOR_SUBHEADER)
FILL_TOTAL = PatternFill("solid", fgColor=COLOR_TOTAL)
FILL_MONTH = PatternFill("solid", fgColor=COLOR_MONTH_ROW)
FILL_TITLE = PatternFill("solid", fgColor=COLOR_TITLE)

# 边框
thin_border = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)

# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def apply_border_range(ws, cell_range):
    """给一个区域的所有单元格加边框"""
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border


def set_column_widths(ws, widths):
    """批量设置列宽 {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ==================== 创建工作簿 ====================
wb = Workbook()

# ============================================================
# Sheet 1: 🛒 购买记录明细
# ============================================================
ws1 = wb.active
ws1.title = "🛒购买记录明细"

# --- 标题行 ---
ws1.merge_cells("A1:I1")
ws1["A1"] = "🛒 购买记录明细表"
ws1["A1"].font = FONT_TITLE
ws1["A1"].fill = FILL_TITLE
ws1["A1"].alignment = ALIGN_CENTER
ws1.row_dimensions[1].height = 32

# --- 表头 ---
headers = ["序号", "购买平台", "下单时间", "商品名称", "规格", "数量", "实付金额(元)", "订单号", "备注"]
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border
ws1.row_dimensions[2].height = 28

# --- 列宽 ---
set_column_widths(ws1, {
    "A": 8, "B": 12, "C": 22, "D": 32, "E": 22,
    "F": 8, "G": 14, "H": 24, "I": 20
})

# --- 预留 1000 行数据区域，设置格式和公式 ---
DATA_START_ROW = 3
DATA_END_ROW = 1002  # 预留 1000 行

for row in range(DATA_START_ROW, DATA_END_ROW + 1):
    # 序号自动公式
    ws1.cell(row=row, column=1, value=f"=IF(C{row}=\"\",\"\",ROW()-2)")
    # 实付金额列设置数字格式
    ws1.cell(row=row, column=7).number_format = "¥#,##0.00"
    # 数量列
    ws1.cell(row=row, column=6).number_format = "0"
    # 所有单元格字体和边框
    for col in range(1, 10):
        c = ws1.cell(row=row, column=col)
        c.font = FONT_NORMAL
        c.border = thin_border
        if col in (1, 2, 3, 6, 7):
            c.alignment = ALIGN_CENTER
        else:
            c.alignment = ALIGN_LEFT
    ws1.row_dimensions[row].height = 22

# --- 合计行（在数据区域下方） ---
TOTAL_ROW = DATA_END_ROW + 2
ws1.merge_cells(f"A{TOTAL_ROW}:F{TOTAL_ROW}")
ws1.cell(row=TOTAL_ROW, column=1, value="📊 合计").font = FONT_TOTAL
ws1.cell(row=TOTAL_ROW, column=1).fill = FILL_TOTAL
ws1.cell(row=TOTAL_ROW, column=1).alignment = ALIGN_CENTER
# 实付金额合计
ws1.cell(row=TOTAL_ROW, column=7, value=f"=SUM(G{DATA_START_ROW}:G{DATA_END_ROW})")
ws1.cell(row=TOTAL_ROW, column=7).font = FONT_TOTAL
ws1.cell(row=TOTAL_ROW, column=7).fill = FILL_TOTAL
ws1.cell(row=TOTAL_ROW, column=7).number_format = "¥#,##0.00"
ws1.cell(row=TOTAL_ROW, column=7).alignment = ALIGN_CENTER
ws1.cell(row=TOTAL_ROW, column=7).border = thin_border
ws1.merge_cells(f"H{TOTAL_ROW}:I{TOTAL_ROW}")
ws1.cell(row=TOTAL_ROW, column=8, value=f"=COUNTA(C{DATA_START_ROW}:C{DATA_END_ROW}) & \" 笔订单\"")
ws1.cell(row=TOTAL_ROW, column=8).font = FONT_TOTAL
ws1.cell(row=TOTAL_ROW, column=8).fill = FILL_TOTAL
ws1.cell(row=TOTAL_ROW, column=8).alignment = ALIGN_CENTER
ws1.cell(row=TOTAL_ROW, column=8).border = thin_border
ws1.row_dimensions[TOTAL_ROW].height = 28

# 冻结表头
ws1.freeze_panes = "A3"

# ============================================================
# Sheet 2: 📊 月度统计
# ============================================================
ws2 = wb.create_sheet("📊月度统计")

ws2.merge_cells("A1:F1")
ws2["A1"] = "📊 月度购买统计"
ws2["A1"].font = FONT_TITLE
ws2["A1"].fill = FILL_TITLE
ws2["A1"].alignment = ALIGN_CENTER
ws2.row_dimensions[1].height = 32

# 表头
month_headers = ["年份", "月份", "购买笔数", "总金额(元)", "平均单价(元)", "占比"]
for col_idx, h in enumerate(month_headers, 1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border
ws2.row_dimensions[2].height = 28

set_column_widths(ws2, {"A": 10, "B": 10, "C": 12, "D": 16, "E": 16, "F": 10})

# 生成近 24 个月（2025-01 到 2026-12）的统计行
months_data = []
years = [2025, 2026]
for y in years:
    for m in range(1, 13):
        months_data.append((y, m))

start_row = 3
for i, (y, m) in enumerate(months_data):
    row = start_row + i
    # 年份
    ws2.cell(row=row, column=1, value=y)
    # 月份
    ws2.cell(row=row, column=2, value=f"{m:02d}月")
    # 购买笔数：COUNTIFS 明细表中下单时间在某年某月
    ws2.cell(row=row, column=3, value=(
        f'=COUNTIFS(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW},">="&DATE({y},{m},1),'
        f'\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW},"<"&DATE({y},{m+1 if m<12 else 1},{1 if m<12 else 1})+IF({m}=12,365,0))'
    ))
    # 简化：直接用 SUMPRODUCT
    ws2.cell(row=row, column=3, value=(
        f'=SUMPRODUCT((YEAR(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW})={y})*'
        f'(MONTH(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW})={m})*'
        f'(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW}<>""))'
    ))
    # 总金额
    ws2.cell(row=row, column=4, value=(
        f'=SUMPRODUCT((YEAR(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW})={y})*'
        f'(MONTH(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW})={m})*'
        f'(\'🛒购买记录明细\'!G{DATA_START_ROW}:G{DATA_END_ROW}))'
    ))
    # 平均单价
    ws2.cell(row=row, column=5, value=f'=IF(C{row}=0,0,D{row}/C{row})')
    # 占比
    ws2.cell(row=row, column=6, value=f'=IF(D{row}=0,0,D{row}/SUM($D${start_row}:$D${start_row+len(months_data)-1}))')

    # 格式
    for col in range(1, 7):
        c = ws2.cell(row=row, column=col)
        c.font = FONT_NORMAL
        c.border = thin_border
        c.alignment = ALIGN_CENTER
    ws2.cell(row=row, column=4).number_format = "¥#,##0.00"
    ws2.cell(row=row, column=5).number_format = "¥#,##0.00"
    ws2.cell(row=row, column=6).number_format = "0.00%"
    ws2.row_dimensions[row].height = 22

# 合计行
month_total_row = start_row + len(months_data)
ws2.merge_cells(f"A{month_total_row}:B{month_total_row}")
ws2.cell(row=month_total_row, column=1, value="📊 合计").font = FONT_TOTAL
ws2.cell(row=month_total_row, column=1).fill = FILL_TOTAL
ws2.cell(row=month_total_row, column=1).alignment = ALIGN_CENTER
ws2.cell(row=month_total_row, column=3, value=f"=SUM(C{start_row}:C{month_total_row-1})")
ws2.cell(row=month_total_row, column=4, value=f"=SUM(D{start_row}:D{month_total_row-1})")
ws2.cell(row=month_total_row, column=5, value=f"=IF(C{month_total_row}=0,0,D{month_total_row}/C{month_total_row})")
ws2.cell(row=month_total_row, column=6, value="100.00%")
for col in range(1, 7):
    c = ws2.cell(row=month_total_row, column=col)
    c.font = FONT_TOTAL
    c.fill = FILL_TOTAL
    c.border = thin_border
    c.alignment = ALIGN_CENTER
ws2.cell(row=month_total_row, column=4).number_format = "¥#,##0.00"
ws2.cell(row=month_total_row, column=5).number_format = "¥#,##0.00"
ws2.row_dimensions[month_total_row].height = 28

# 月度金额柱状图
bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.style = 10
bar_chart.title = "月度购买金额趋势"
bar_chart.y_axis.title = "金额(元)"
bar_chart.x_axis.title = "月份"
data_ref = Reference(ws2, min_col=4, min_row=2, max_row=month_total_row-1, max_col=4)
cats_ref = Reference(ws2, min_col=2, min_row=3, max_row=month_total_row-1)
bar_chart.add_data(data_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)
bar_chart.height = 10
bar_chart.width = 22
ws2.add_chart(bar_chart, f"H2")

# 月度笔数折线图
line_chart = LineChart()
line_chart.title = "月度购买笔数趋势"
line_chart.y_axis.title = "笔数"
line_chart.x_axis.title = "月份"
data_ref2 = Reference(ws2, min_col=3, min_row=2, max_row=month_total_row-1, max_col=3)
line_chart.add_data(data_ref2, titles_from_data=True)
line_chart.set_categories(cats_ref)
line_chart.height = 10
line_chart.width = 22
ws2.add_chart(line_chart, f"H22")

ws2.freeze_panes = "A3"

# ============================================================
# Sheet 3: 📈 年度统计
# ============================================================
ws3 = wb.create_sheet("📈年度统计")

ws3.merge_cells("A1:E1")
ws3["A1"] = "📈 年度购买统计"
ws3["A1"].font = FONT_TITLE
ws3["A1"].fill = FILL_TITLE
ws3["A1"].alignment = ALIGN_CENTER
ws3.row_dimensions[1].height = 32

year_headers = ["年份", "购买笔数", "总金额(元)", "平均单价(元)", "占比"]
for col_idx, h in enumerate(year_headers, 1):
    cell = ws3.cell(row=2, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border
ws3.row_dimensions[2].height = 28

set_column_widths(ws3, {"A": 10, "B": 12, "C": 16, "D": 16, "E": 10})

year_start = 3
for i, y in enumerate(years):
    row = year_start + i
    ws3.cell(row=row, column=1, value=y)
    ws3.cell(row=row, column=2, value=(
        f'=SUMPRODUCT((YEAR(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW})={y})*'
        f'(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW}<>""))'
    ))
    ws3.cell(row=row, column=3, value=(
        f'=SUMPRODUCT((YEAR(\'🛒购买记录明细\'!C{DATA_START_ROW}:C{DATA_END_ROW})={y})*'
        f'(\'🛒购买记录明细\'!G{DATA_START_ROW}:G{DATA_END_ROW}))'
    ))
    ws3.cell(row=row, column=4, value=f'=IF(B{row}=0,0,C{row}/B{row})')
    ws3.cell(row=row, column=5, value=f'=IF(C{row}=0,0,C{row}/SUM($C${year_start}:$C${year_start+len(years)-1}))')
    for col in range(1, 6):
        c = ws3.cell(row=row, column=col)
        c.font = FONT_NORMAL
        c.border = thin_border
        c.alignment = ALIGN_CENTER
    ws3.cell(row=row, column=3).number_format = "¥#,##0.00"
    ws3.cell(row=row, column=4).number_format = "¥#,##0.00"
    ws3.cell(row=row, column=5).number_format = "0.00%"
    ws3.row_dimensions[row].height = 24

year_total_row = year_start + len(years)
ws3.cell(row=year_total_row, column=1, value="📊 合计").font = FONT_TOTAL
ws3.cell(row=year_total_row, column=1).fill = FILL_TOTAL
ws3.cell(row=year_total_row, column=2, value=f"=SUM(B{year_start}:B{year_total_row-1})")
ws3.cell(row=year_total_row, column=3, value=f"=SUM(C{year_start}:C{year_total_row-1})")
ws3.cell(row=year_total_row, column=4, value=f"=IF(B{year_total_row}=0,0,C{year_total_row}/B{year_total_row})")
ws3.cell(row=year_total_row, column=5, value="100.00%")
for col in range(1, 6):
    c = ws3.cell(row=year_total_row, column=col)
    c.font = FONT_TOTAL
    c.fill = FILL_TOTAL
    c.border = thin_border
    c.alignment = ALIGN_CENTER
ws3.cell(row=year_total_row, column=3).number_format = "¥#,##0.00"
ws3.cell(row=year_total_row, column=4).number_format = "¥#,##0.00"
ws3.row_dimensions[year_total_row].height = 28

# 年度柱状图
year_bar = BarChart()
year_bar.type = "col"
year_bar.style = 12
year_bar.title = "年度购买金额对比"
year_bar.y_axis.title = "金额(元)"
year_bar.x_axis.title = "年份"
y_data = Reference(ws3, min_col=3, min_row=2, max_row=year_total_row-1, max_col=3)
y_cats = Reference(ws3, min_col=1, min_row=3, max_row=year_total_row-1)
year_bar.add_data(y_data, titles_from_data=True)
year_bar.set_categories(y_cats)
year_bar.height = 10
year_bar.width = 18
year_bar.dataLabels = DataLabelList(showVal=True)
ws3.add_chart(year_bar, "G2")

ws3.freeze_panes = "A3"

# ============================================================
# Sheet 4: 🏪 平台分布
# ============================================================
ws4 = wb.create_sheet("🏪平台分布")

ws4.merge_cells("A1:D1")
ws4["A1"] = "🏪 购买平台分布"
ws4["A1"].font = FONT_TITLE
ws4["A1"].fill = FILL_TITLE
ws4["A1"].alignment = ALIGN_CENTER
ws4.row_dimensions[1].height = 32

platform_headers = ["购买平台", "购买笔数", "总金额(元)", "占比"]
for col_idx, h in enumerate(platform_headers, 1):
    cell = ws4.cell(row=2, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border
ws4.row_dimensions[2].height = 28

set_column_widths(ws4, {"A": 14, "B": 12, "C": 16, "D": 10})

platforms = ["抖音", "1688", "淘宝", "京东", "拼多多", "天猫", "苏宁", "其他"]
plat_start = 3
for i, p in enumerate(platforms):
    row = plat_start + i
    ws4.cell(row=row, column=1, value=p)
    ws4.cell(row=row, column=2, value=(
        f'=COUNTIF(\'🛒购买记录明细\'!B{DATA_START_ROW}:B{DATA_END_ROW},"{p}")'
    ))
    ws4.cell(row=row, column=3, value=(
        f'=SUMIF(\'🛒购买记录明细\'!B{DATA_START_ROW}:B{DATA_END_ROW},"{p}",'
        f'\'🛒购买记录明细\'!G{DATA_START_ROW}:G{DATA_END_ROW})'
    ))
    ws4.cell(row=row, column=4, value=f'=IF(C{row}=0,0,C{row}/SUM($C${plat_start}:$C${plat_start+len(platforms)-1}))')
    for col in range(1, 5):
        c = ws4.cell(row=row, column=col)
        c.font = FONT_NORMAL
        c.border = thin_border
        c.alignment = ALIGN_CENTER
    ws4.cell(row=row, column=3).number_format = "¥#,##0.00"
    ws4.cell(row=row, column=4).number_format = "0.00%"
    ws4.row_dimensions[row].height = 22

plat_total_row = plat_start + len(platforms)
ws4.cell(row=plat_total_row, column=1, value="📊 合计").font = FONT_TOTAL
ws4.cell(row=plat_total_row, column=1).fill = FILL_TOTAL
ws4.cell(row=plat_total_row, column=2, value=f"=SUM(B{plat_start}:B{plat_total_row-1})")
ws4.cell(row=plat_total_row, column=3, value=f"=SUM(C{plat_start}:C{plat_total_row-1})")
ws4.cell(row=plat_total_row, column=4, value="100.00%")
for col in range(1, 5):
    c = ws4.cell(row=plat_total_row, column=col)
    c.font = FONT_TOTAL
    c.fill = FILL_TOTAL
    c.border = thin_border
    c.alignment = ALIGN_CENTER
ws4.cell(row=plat_total_row, column=3).number_format = "¥#,##0.00"
ws4.row_dimensions[plat_total_row].height = 28

# 平台金额饼图
pie = PieChart()
pie.title = "各平台购买金额占比"
pie_data = Reference(ws4, min_col=3, min_row=2, max_row=plat_total_row-1, max_col=3)
pie_cats = Reference(ws4, min_col=1, min_row=3, max_row=plat_total_row-1)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.height = 10
pie.width = 14
pie.dataLabels = DataLabelList(showPercent=True)
ws4.add_chart(pie, "F2")

# 平台笔数柱状图
plat_bar = BarChart()
plat_bar.type = "bar"
plat_bar.style = 11
plat_bar.title = "各平台购买笔数"
plat_bar.y_axis.title = "平台"
plat_bar.x_axis.title = "笔数"
pb_data = Reference(ws4, min_col=2, min_row=2, max_row=plat_total_row-1, max_col=2)
pb_cats = Reference(ws4, min_col=1, min_row=3, max_row=plat_total_row-1)
plat_bar.add_data(pb_data, titles_from_data=True)
plat_bar.set_categories(pb_cats)
plat_bar.height = 10
plat_bar.width = 14
ws4.add_chart(plat_bar, "F22")

ws4.freeze_panes = "A3"

# ============================================================
# 保存
# ============================================================
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"✅ 已生成: {OUTPUT_FILE}")
print(f"   包含 4 个工作表:")
print(f"   1. 🛒购买记录明细 - 详细记录(预留1000行)")
print(f"   2. 📊月度统计 - 含柱状图+折线图")
print(f"   3. 📈年度统计 - 含柱状图")
print(f"   4. 🏪平台分布 - 含饼图+柱状图")
