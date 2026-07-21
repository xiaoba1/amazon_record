"""
购买记录归档脚本（含出货记录绑定）
功能：
  1. 写入明细记录（序号用真实数字）
  2. 写入出货记录并与购买记录关联
  3. 自动重算并写入 月度统计/年度统计/平台分布 的数值
  4. 自动 git commit + push

字段：序号 / 购买平台 / 下单时间 / 商品名称 / 规格 / 数量 / 实付金额 / 订单号 / 关联出货单 / 备注
"""
import os
import subprocess
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

XLSX = "/workspace/购买记录/归档表格/购买记录_2026.xlsx"
YEAR = 2026

# ==================== 样式 ====================
COLOR_HEADER = "2F5496"
COLOR_TOTAL = "FFF2CC"
COLOR_BORDER = "8EAADB"
COLOR_TITLE = "1F4E79"
FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="微软雅黑", size=10, color="333333")
FONT_TOTAL = Font(name="微软雅黑", size=11, bold=True, color="C00000")
FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_TOTAL = PatternFill("solid", fgColor=COLOR_TOTAL)
FILL_TITLE = PatternFill("solid", fgColor=COLOR_TITLE)
thin_border = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def auto_fit_columns(ws, headers, data_start_row=3, data_end_row=502, max_width=50, min_width=8):
    """自适应列宽：根据表头+数据内容计算每列最大宽度
    日元列(出货表9-12)显示为 'JPY 1,234'，元列(购买表7)显示为 '¥1,234.00'
    """
    from openpyxl.utils import get_column_letter
    # 判断是否为出货记录表（表头含"销售价(日元)"）
    is_shipment = any("日元" in str(h) for h in headers)
    jpy_cols = (9, 10, 11, 12) if is_shipment else ()
    cny_cols = (7,) if not is_shipment else ()

    num_cols = len(headers)
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1] or ""))
        for row in range(data_start_row, min(data_end_row + 1, data_start_row + 200)):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            if isinstance(val, (int, float)):
                if col_idx in jpy_cols:
                    s = f"JPY {int(val):,}"
                elif col_idx in cny_cols:
                    s = f"¥{val:,.2f}"
                else:
                    s = str(val)
            else:
                s = str(val)
            width = sum(2 if ord(c) > 127 else 1 for c in s)
            if width > max_len:
                max_len = width
        final_width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = final_width

# 字段顺序（无截图列，增加关联出货单列）
HEADERS = ["序号", "购买平台", "下单时间", "商品名称", "规格", "购买数量", "实付金额(元)", "购买订单号", "关联出货单", "出货数量", "库存", "备注"]
NUM_COLS = len(HEADERS)  # 12

# 出货记录字段
SHIPMENT_HEADERS = ["序号", "出货平台", "订购日期", "发货日期", "商品名称", "规格", "SKU", "数量", "销售价(日元)", "税金(日元)", "手续费(日元)", "销售额(日元)", "出货订单号", "关联购买记录"]
SHIPMENT_NUM_COLS = len(SHIPMENT_HEADERS)  # 14


def auto_match_stock(ship_record, purchase_records):
    """智能匹配有库存的购买记录
    
    匹配规则：
    1. 库存 > 0
    2. 商品名称相似度匹配（关键词匹配）
    
    返回：匹配到的购买订单号，未匹配返回 None
    """
    from difflib import SequenceMatcher
    
    ship_product = str(ship_record.get("product", "")).lower()
    ship_spec = str(ship_record.get("spec", "")).lower()
    
    candidates = []
    
    for rec in purchase_records:
        stock = int(rec.get("stock", 0) or 0)
        if stock <= 0:
            continue  # 跳过无库存
        
        pur_product = str(rec.get("product", "")).lower()
        pur_spec = str(rec.get("spec", "")).lower()
        
        # 计算相似度
        product_ratio = SequenceMatcher(None, ship_product, pur_product).ratio()
        spec_ratio = SequenceMatcher(None, ship_spec, pur_spec).ratio() if ship_spec and pur_spec else 0
        
        # 关键词匹配（包含关系）
        keyword_match = False
        if ship_product and pur_product:
            # 提取关键词（去除常见词）
            keywords = set(ship_product.replace("玻璃", "").replace("修复", "").replace("膏", "").split())
            pur_keywords = set(pur_product.replace("玻璃", "").replace("修复", "").replace("膏", "").split())
            if keywords & pur_keywords:
                keyword_match = True
        
        # 综合得分
        score = product_ratio * 0.5 + spec_ratio * 0.3 + (0.2 if keyword_match else 0)
        
        if score > 0.3:  # 阈值
            candidates.append((score, rec["order"], rec.get("product", ""), stock))
    
    if candidates:
        # 按得分降序排序，选择最高分
        candidates.sort(reverse=True)
        best = candidates[0]
        print(f"    [匹配候选] 得分={best[0]:.2f} | 商品={best[2]} | 库存={best[3]}")
        return best[1]
    
    return None


def get_existing_records(ws):
    """从明细表读取所有已有记录"""
    records = []
    row = 3
    while row <= 502:
        time_val = ws.cell(row=row, column=3).value
        if not time_val:
            break
        records.append({
            "platform": ws.cell(row=row, column=2).value or "",
            "time": time_val,
            "product": ws.cell(row=row, column=4).value or "",
            "spec": ws.cell(row=row, column=5).value or "",
            "qty": ws.cell(row=row, column=6).value or 0,
            "paid": ws.cell(row=row, column=7).value or 0,
            "order": ws.cell(row=row, column=8).value or "",
            "shipment": ws.cell(row=row, column=9).value or "",
            "ship_qty": ws.cell(row=row, column=10).value or 0,
            "stock": ws.cell(row=row, column=11).value or 0,
            "remark": ws.cell(row=row, column=12).value or "",
        })
        row += 1
    return records


def get_existing_shipments(ws):
    """从出货记录表读取所有已有出货记录"""
    shipments = []
    row = 3
    while row <= 502:
        order_val = ws.cell(row=row, column=13).value
        if not order_val:
            break
        shipments.append({
            "ship_platform": ws.cell(row=row, column=2).value or "",
            "order_date": ws.cell(row=row, column=3).value,
            "ship_date": ws.cell(row=row, column=4).value,
            "product": ws.cell(row=row, column=5).value or "",
            "spec": ws.cell(row=row, column=6).value or "",
            "sku": ws.cell(row=row, column=7).value or "",
            "qty": ws.cell(row=row, column=8).value or 0,
            "price": ws.cell(row=row, column=9).value or 0,
            "tax": ws.cell(row=row, column=10).value or 0,
            "fee": ws.cell(row=row, column=11).value or 0,
            "revenue": ws.cell(row=row, column=12).value or 0,
            "ship_order": ws.cell(row=row, column=13).value or "",
            "purchase_ref": ws.cell(row=row, column=14).value or "",
        })
        row += 1
    return shipments


def unmerge_all(ws):
    """解除工作表所有合并单元格"""
    merged = list(ws.merged_cells.ranges)
    for r in merged:
        ws.unmerge_cells(str(r))


def rewrite_detail_sheet(ws, records, shipment_row_map=None):
    """重写明细表
    shipment_row_map: {出货单号: 出货记录表行号}，用于建立超链接
    """
    unmerge_all(ws)
    # 清空数据区（含标题行、表头行）
    for row in range(1, 503):
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=row, column=col).value = None

    # 第1行：标题（合并单元格）
    ws.merge_cells(f"A1:{chr(64 + NUM_COLS)}1")
    title_cell = ws.cell(row=1, column=1, value=f"🛒 购买记录明细表 ({YEAR}年)")
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    # 第2行：表头
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 数据行
    for i, rec in enumerate(records):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1)  # 真实序号
        ws.cell(row=row, column=2, value=rec["platform"])
        ws.cell(row=row, column=3, value=rec["time"])
        ws.cell(row=row, column=3).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row, column=4, value=rec["product"])
        ws.cell(row=row, column=5, value=rec["spec"])
        ws.cell(row=row, column=6, value=rec["qty"])
        ws.cell(row=row, column=7, value=rec["paid"])
        ws.cell(row=row, column=8, value=rec["order"])
        # 关联出货单：写入带超链接的文本（点击跳转到出货记录表对应行）
        # 支持多个出货单（逗号分隔），但只有第一个设置超链接
        if rec.get("shipment"):
            shipment_str = str(rec["shipment"])
            ship_cell = ws.cell(row=row, column=9, value=shipment_str)
            
            # 如果是单个出货单，设置超链接
            if "," not in shipment_str and shipment_row_map:
                ship_row = shipment_row_map.get(shipment_str)
                if ship_row:
                    from openpyxl.worksheet.hyperlink import Hyperlink
                    ship_cell.hyperlink = Hyperlink(ref=f"I{row}", location=f"'📦出货记录'!A{ship_row}", display=shipment_str)
                    ship_cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
                else:
                    ship_cell.font = FONT_NORMAL
            else:
                # 多个出货单，只显示文本（Excel 限制单单元格只能一个超链接）
                ship_cell.font = FONT_NORMAL
            ship_cell.alignment = ALIGN_CENTER
        # 出货数量和库存
        ws.cell(row=row, column=10, value=rec.get("ship_qty", 0))
        ws.cell(row=row, column=11, value=rec.get("stock", 0))
        if rec.get("remark"):
            ws.cell(row=row, column=12, value=rec["remark"])

        for col in range(1, NUM_COLS + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if col == 9 and rec.get("shipment"):
                continue
            c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER if col in (1, 2, 3, 6, 7, 10, 11) else ALIGN_LEFT
        ws.cell(row=row, column=7).number_format = "¥#,##0.00"
        ws.cell(row=row, column=6).number_format = "0"
        ws.cell(row=row, column=10).number_format = "0"
        ws.cell(row=row, column=11).number_format = "0"
        ws.row_dimensions[row].height = 22

    # 合计行
    if records:
        total_row = 3 + len(records) + 1
        ws.merge_cells(f"A{total_row}:F{total_row}")
        ws.cell(row=total_row, column=1, value="📊 合计")
        ws.cell(row=total_row, column=1).font = FONT_TOTAL
        ws.cell(row=total_row, column=1).fill = FILL_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        total_paid = sum(float(r["paid"] or 0) for r in records)
        ws.cell(row=total_row, column=7, value=round(total_paid, 2))
        ws.cell(row=total_row, column=7).font = FONT_TOTAL
        ws.cell(row=total_row, column=7).fill = FILL_TOTAL
        ws.cell(row=total_row, column=7).number_format = "¥#,##0.00"
        ws.cell(row=total_row, column=7).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=7).border = thin_border
        ws.merge_cells(f"H{total_row}:J{total_row}")
        ws.cell(row=total_row, column=8, value=f"{len(records)} 笔订单")
        ws.cell(row=total_row, column=8).font = FONT_TOTAL
        ws.cell(row=total_row, column=8).fill = FILL_TOTAL
        ws.cell(row=total_row, column=8).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=8).border = thin_border
        # 库存统计
        total_stock = sum(float(r.get("stock", 0) or 0) for r in records)
        ws.cell(row=total_row, column=11, value=int(total_stock))
        ws.cell(row=total_row, column=11).font = FONT_TOTAL
        ws.cell(row=total_row, column=11).fill = FILL_TOTAL
        ws.cell(row=total_row, column=11).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=11).border = thin_border
        ws.cell(row=total_row, column=12, value="库存合计")
        ws.cell(row=total_row, column=12).font = FONT_TOTAL
        ws.cell(row=total_row, column=12).fill = FILL_TOTAL
        ws.cell(row=total_row, column=12).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=12).border = thin_border
        ws.row_dimensions[total_row].height = 28

    # 自适应列宽
    auto_fit_columns(ws, HEADERS)
    ws.freeze_panes = "A3"


def rewrite_shipment_sheet(ws, shipments, purchase_row_map=None):
    """重写出货记录表
    purchase_row_map: {购买订单号: 购买记录表行号}，用于建立超链接
    """
    unmerge_all(ws)
    # 清空数据区（含标题行、表头行）
    for row in range(1, 503):
        for col in range(1, SHIPMENT_NUM_COLS + 1):
            ws.cell(row=row, column=col).value = None

    # 第1行：标题（合并单元格）
    last_col_letter = chr(64 + SHIPMENT_NUM_COLS) if SHIPMENT_NUM_COLS <= 26 else "N"
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=1, value=f"📦 出货记录表 ({YEAR}年)")
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    # 第2行：表头
    for col_idx, h in enumerate(SHIPMENT_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 数据行
    for i, ship in enumerate(shipments):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1)  # 真实序号
        ws.cell(row=row, column=2, value=ship["ship_platform"])
        if ship["order_date"]:
            ws.cell(row=row, column=3, value=ship["order_date"])
            ws.cell(row=row, column=3).number_format = "yyyy-mm-dd"
        if ship["ship_date"]:
            ws.cell(row=row, column=4, value=ship["ship_date"])
            ws.cell(row=row, column=4).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=5, value=ship["product"])
        ws.cell(row=row, column=6, value=ship["spec"])
        ws.cell(row=row, column=7, value=ship["sku"])
        ws.cell(row=row, column=8, value=ship["qty"])
        ws.cell(row=row, column=9, value=ship["price"])
        ws.cell(row=row, column=10, value=ship["tax"])
        ws.cell(row=row, column=11, value=ship["fee"])
        ws.cell(row=row, column=12, value=ship["revenue"])
        ws.cell(row=row, column=13, value=ship["ship_order"])
        # 关联购买记录：写入带超链接的文本（点击跳转到购买记录明细表对应行）
        if ship.get("purchase_ref"):
            pur_order = str(ship["purchase_ref"])
            pur_row = purchase_row_map.get(pur_order) if purchase_row_map else None
            pur_cell = ws.cell(row=row, column=14, value=pur_order)
            if pur_row:
                # 工作簿内部跳转：必须用 Hyperlink 对象 + location 属性
                from openpyxl.worksheet.hyperlink import Hyperlink
                pur_cell.hyperlink = Hyperlink(ref=f"N{row}", location=f"'🛒购买记录明细'!A{pur_row}", display=pur_order)
                pur_cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
            else:
                pur_cell.font = FONT_NORMAL
            pur_cell.alignment = ALIGN_CENTER

        for col in range(1, SHIPMENT_NUM_COLS + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if col == 14 and ship.get("purchase_ref") and purchase_row_map and pur_row:
                continue  # 已单独设置字体
            c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER if col in (1, 2, 3, 4, 7, 8, 9, 10, 11, 12, 13) else ALIGN_LEFT
        ws.cell(row=row, column=9).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=10).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=11).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=12).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=8).number_format = "0"
        ws.row_dimensions[row].height = 22

    # 合计行
    if shipments:
        total_row = 3 + len(shipments) + 1
        ws.merge_cells(f"A{total_row}:D{total_row}")
        ws.cell(row=total_row, column=1, value="📊 合计")
        ws.cell(row=total_row, column=1).font = FONT_TOTAL
        ws.cell(row=total_row, column=1).fill = FILL_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        total_qty = sum(float(s["qty"] or 0) for s in shipments)
        total_revenue = sum(float(s["revenue"] or 0) for s in shipments)
        ws.cell(row=total_row, column=8, value=int(total_qty))
        ws.cell(row=total_row, column=8).font = FONT_TOTAL
        ws.cell(row=total_row, column=8).fill = FILL_TOTAL
        ws.cell(row=total_row, column=8).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=8).border = thin_border
        ws.merge_cells(f"I{total_row}:L{total_row}")
        ws.cell(row=total_row, column=9, value=f"总销售额: JPY {int(total_revenue):,}")
        ws.cell(row=total_row, column=9).font = FONT_TOTAL
        ws.cell(row=total_row, column=9).fill = FILL_TOTAL
        ws.cell(row=total_row, column=9).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=9).border = thin_border
        ws.merge_cells(f"M{total_row}:N{total_row}")
        ws.cell(row=total_row, column=13, value=f"{len(shipments)} 笔出货")
        ws.cell(row=total_row, column=13).font = FONT_TOTAL
        ws.cell(row=total_row, column=13).fill = FILL_TOTAL
        ws.cell(row=total_row, column=13).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=13).border = thin_border
        ws.row_dimensions[total_row].height = 28

    # 自适应列宽
    auto_fit_columns(ws, SHIPMENT_HEADERS)
    ws.freeze_panes = "A3"


def rewrite_month_sheet(ws, records):
    """重写月度统计"""
    unmerge_all(ws)
    for row in range(3, 20):
        for col in range(1, 7):
            ws.cell(row=row, column=col).value = None

    month_stats = defaultdict(lambda: {"count": 0, "amount": 0})
    for r in records:
        t = r["time"]
        if not t:
            continue
        t = t if isinstance(t, datetime) else datetime.fromisoformat(str(t))
        if t.year == YEAR:
            month_stats[t.month]["count"] += 1
            month_stats[t.month]["amount"] += float(r["paid"] or 0)

    total_amount = sum(s["amount"] for s in month_stats.values())

    for m in range(1, 13):
        row = 2 + m
        s = month_stats[m]
        ws.cell(row=row, column=1, value=YEAR)
        ws.cell(row=row, column=2, value=f"{m:02d}月")
        ws.cell(row=row, column=3, value=s["count"])
        ws.cell(row=row, column=4, value=round(s["amount"], 2))
        ws.cell(row=row, column=5, value=round(s["amount"] / s["count"], 2) if s["count"] else 0)
        ws.cell(row=row, column=6, value=round(s["amount"] / total_amount, 4) if total_amount else 0)
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.font = FONT_NORMAL
            c.border = thin_border
            c.alignment = ALIGN_CENTER
        ws.cell(row=row, column=4).number_format = "¥#,##0.00"
        ws.cell(row=row, column=5).number_format = "¥#,##0.00"
        ws.cell(row=row, column=6).number_format = "0.00%"

    total_row = 15
    total_count = sum(s["count"] for s in month_stats.values())
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws.cell(row=total_row, column=1, value="📊 合计")
    ws.cell(row=total_row, column=3, value=total_count)
    ws.cell(row=total_row, column=4, value=round(total_amount, 2))
    ws.cell(row=total_row, column=5, value=round(total_amount / total_count, 2) if total_count else 0)
    ws.cell(row=total_row, column=6, value=1.0)
    for col in range(1, 7):
        c = ws.cell(row=total_row, column=col)
        c.font = FONT_TOTAL
        c.fill = FILL_TOTAL
        c.border = thin_border
        c.alignment = ALIGN_CENTER
    ws.cell(row=total_row, column=4).number_format = "¥#,##0.00"
    ws.cell(row=total_row, column=5).number_format = "¥#,##0.00"
    ws.cell(row=total_row, column=6).number_format = "0.00%"

    # 图表
    ws._charts = []
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = f"{YEAR}年 月度购买金额趋势"
    bar.y_axis.title = "金额(元)"
    bar.x_axis.title = "月份"
    data = Reference(ws, min_col=4, min_row=2, max_row=14, max_col=4)
    cats = Reference(ws, min_col=2, min_row=3, max_row=14)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 10
    bar.width = 22
    ws.add_chart(bar, "H2")

    line = LineChart()
    line.title = f"{YEAR}年 月度购买笔数趋势"
    line.y_axis.title = "笔数"
    line.x_axis.title = "月份"
    data2 = Reference(ws, min_col=3, min_row=2, max_row=14, max_col=3)
    line.add_data(data2, titles_from_data=True)
    line.set_categories(cats)
    line.height = 10
    line.width = 22
    ws.add_chart(line, "H22")


def rewrite_year_sheet(ws, records):
    """重写年度统计"""
    unmerge_all(ws)
    for row in range(3, 10):
        for col in range(1, 6):
            ws.cell(row=row, column=col).value = None

    year_count = 0
    year_amount = 0
    for r in records:
        t = r["time"]
        if not t:
            continue
        t = t if isinstance(t, datetime) else datetime.fromisoformat(str(t))
        if t.year == YEAR:
            year_count += 1
            year_amount += float(r["paid"] or 0)

    row = 3
    ws.cell(row=row, column=1, value=YEAR)
    ws.cell(row=row, column=2, value=year_count)
    ws.cell(row=row, column=3, value=round(year_amount, 2))
    ws.cell(row=row, column=4, value=round(year_amount / year_count, 2) if year_count else 0)
    for col in range(1, 6):
        c = ws.cell(row=row, column=col)
        c.font = FONT_NORMAL
        c.border = thin_border
        c.alignment = ALIGN_CENTER
    ws.cell(row=row, column=3).number_format = "¥#,##0.00"
    ws.cell(row=row, column=4).number_format = "¥#,##0.00"

    ws._charts = []
    bar = BarChart()
    bar.type = "col"
    bar.style = 12
    bar.title = f"{YEAR}年 购买统计"
    bar.y_axis.title = "金额(元)"
    data = Reference(ws, min_col=3, min_row=2, max_row=3, max_col=3)
    cats = Reference(ws, min_col=1, min_row=3, max_row=3)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 14
    bar.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(bar, "G2")


def rewrite_platform_sheet(ws, records):
    """重写平台分布"""
    unmerge_all(ws)
    for row in range(3, 20):
        for col in range(1, 5):
            ws.cell(row=row, column=col).value = None

    platforms = ["抖音", "抖音商城", "1688", "淘宝", "京东", "拼多多", "天猫", "苏宁", "其他"]
    plat_stats = defaultdict(lambda: {"count": 0, "amount": 0})
    for r in records:
        t = r["time"]
        if not t:
            continue
        t = t if isinstance(t, datetime) else datetime.fromisoformat(str(t))
        if t.year != YEAR:
            continue
        p = r["platform"] if r["platform"] in platforms else "其他"
        plat_stats[p]["count"] += 1
        plat_stats[p]["amount"] += float(r["paid"] or 0)

    total_amount = sum(s["amount"] for s in plat_stats.values())
    plat_start = 3
    for i, p in enumerate(platforms):
        row = plat_start + i
        s = plat_stats[p]
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=2, value=s["count"])
        ws.cell(row=row, column=3, value=round(s["amount"], 2))
        ws.cell(row=row, column=4, value=round(s["amount"] / total_amount, 4) if total_amount else 0)
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            c.font = FONT_NORMAL
            c.border = thin_border
            c.alignment = ALIGN_CENTER
        ws.cell(row=row, column=3).number_format = "¥#,##0.00"
        ws.cell(row=row, column=4).number_format = "0.00%"

    plat_total_row = plat_start + len(platforms)
    total_count = sum(s["count"] for s in plat_stats.values())
    ws.cell(row=plat_total_row, column=1, value="📊 合计")
    ws.cell(row=plat_total_row, column=2, value=total_count)
    ws.cell(row=plat_total_row, column=3, value=round(total_amount, 2))
    ws.cell(row=plat_total_row, column=4, value=1.0 if total_amount else 0)
    for col in range(1, 5):
        c = ws.cell(row=plat_total_row, column=col)
        c.font = FONT_TOTAL
        c.fill = FILL_TOTAL
        c.border = thin_border
        c.alignment = ALIGN_CENTER
    ws.cell(row=plat_total_row, column=3).number_format = "¥#,##0.00"
    ws.cell(row=plat_total_row, column=4).number_format = "0.00%"

    ws._charts = []
    pie = PieChart()
    pie.title = f"{YEAR}年 各平台购买金额占比"
    pie_data = Reference(ws, min_col=3, min_row=2, max_row=plat_total_row - 1, max_col=3)
    pie_cats = Reference(ws, min_col=1, min_row=3, max_row=plat_total_row - 1)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.height = 10
    pie.width = 14
    pie.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(pie, "F2")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 11
    bar.title = f"{YEAR}年 各平台购买笔数"
    bar.y_axis.title = "平台"
    bar.x_axis.title = "笔数"
    bdata = Reference(ws, min_col=2, min_row=2, max_row=plat_total_row - 1, max_col=2)
    bcats = Reference(ws, min_col=1, min_row=3, max_row=plat_total_row - 1)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    bar.height = 10
    bar.width = 14
    ws.add_chart(bar, "F22")


def add_records(new_records, new_shipments=None):
    """添加新记录并重算所有统计"""
    wb = load_workbook(XLSX)
    ws_detail = wb["🛒购买记录明细"]

    existing = get_existing_records(ws_detail)
    print(f"已有购买记录: {len(existing)} 条")

    for rec in new_records:
        existing.append(rec)
        print(f"  + 新增购买: {rec['platform']} | {rec['time']} | {rec['product']} | ¥{rec['paid']}")

    # 按订单号去重：新记录替换旧记录（保留最新的）
    seen_orders = {}
    deduplicated = []
    for rec in existing:
        if rec["order"] in seen_orders:
            print(f"  ↻ 更新购买: {rec['platform']} | {rec['order']}")
        seen_orders[rec["order"]] = rec
    deduplicated = list(seen_orders.values())
    deduplicated.sort(key=lambda x: x["time"])
    print(f"  去重后: {len(deduplicated)} 条")

    # 先处理出货记录（包括新增），便于反向回填购买记录的"关联出货单"
    existing_shipments = []
    if "📦出货记录" in wb.sheetnames:
        ws_shipment = wb["📦出货记录"]
        existing_shipments = get_existing_shipments(ws_shipment)
        print(f"已有出货记录: {len(existing_shipments)} 条")

        if new_shipments:
            for ship in new_shipments:
                # ===== 智能库存匹配 =====
                # 如果没有指定 purchase_ref，尝试自动匹配有库存的购买记录
                if not ship.get("purchase_ref"):
                    matched_order = auto_match_stock(ship, deduplicated)
                    if matched_order:
                        ship["purchase_ref"] = matched_order
                        print(f"  🔗 自动匹配库存: 出货单 {ship['ship_order']} → 购买单 {matched_order}")
                
                existing_shipments.append(ship)
                print(f"  + 新增出货: {ship['ship_platform']} | {ship['order_date']} | {ship['product']} | ¥{ship['revenue']}")

        # ===== 更新购买记录的出货数量和库存 =====
        # 统计每个购买订单的出货数量
        purchase_ship_qty = {}
        for s in existing_shipments:
            if s.get("purchase_ref"):
                pur_order = s["purchase_ref"]
                purchase_ship_qty[pur_order] = purchase_ship_qty.get(pur_order, 0) + int(s.get("qty", 1))
        
        # 更新购买记录
        for rec in deduplicated:
            if rec["order"] in purchase_ship_qty:
                old_ship_qty = rec.get("ship_qty", 0)
                new_ship_qty = purchase_ship_qty[rec["order"]]
                if old_ship_qty != new_ship_qty:
                    rec["ship_qty"] = new_ship_qty
                    rec["stock"] = int(rec.get("qty", 0)) - new_ship_qty
                    print(f"  📦 更新库存: {rec['order']} | 出货 {new_ship_qty} | 库存 {rec['stock']}")

    # 反向回填：根据出货记录的 purchase_ref，把对应出货单号写到购买记录的 shipment 字段
    # 支持一个购买记录对应多个出货记录（用逗号分隔）
    pur_to_ships = {}
    for s in existing_shipments:
        if s.get("purchase_ref"):
            pur_order = s["purchase_ref"]
            if pur_order not in pur_to_ships:
                pur_to_ships[pur_order] = []
            pur_to_ships[pur_order].append(s["ship_order"])
    
    filled = 0
    for rec in deduplicated:
        if rec["order"] in pur_to_ships:
            ships = pur_to_ships[rec["order"]]
            rec["shipment"] = ", ".join(ships)  # 多个出货单用逗号分隔
            filled += 1
    if filled:
        print(f"  ↻ 回填关联出货单: {filled} 条")

    # 构建行号映射（用于超链接互跳）
    # 购买记录：第3行开始，每条记录占1行 -> {购买订单号: 行号}
    purchase_row_map = {rec["order"]: 3 + i for i, rec in enumerate(deduplicated)}
    # 出货记录：第3行开始，每条记录占1行 -> {出货单号: 行号}
    shipment_row_map = {s["ship_order"]: 3 + i for i, s in enumerate(existing_shipments)}

    # 重写购买记录明细（带出货单超链接）
    rewrite_detail_sheet(ws_detail, deduplicated, shipment_row_map)

    # 重写出货记录（带购买记录超链接）
    if "📦出货记录" in wb.sheetnames:
        rewrite_shipment_sheet(ws_shipment, existing_shipments, purchase_row_map)
        # 调整 sheet 顺序：出货记录紧跟在购买记录明细之后
        desired_idx = wb.sheetnames.index("🛒购买记录明细") + 1
        cur_idx = wb.sheetnames.index("📦出货记录")
        if cur_idx != desired_idx:
            wb.move_sheet("📦出货记录", offset=desired_idx - cur_idx)

    rewrite_month_sheet(wb["📊月度统计"], deduplicated)
    rewrite_year_sheet(wb["📈年度统计"], deduplicated)
    rewrite_platform_sheet(wb["🏪平台分布"], deduplicated)

    wb.save(XLSX)
    print(f"\n✅ 已保存，购买记录共 {len(deduplicated)} 条")
    if "📦出货记录" in wb.sheetnames:
        print(f"   出货记录共 {len(existing_shipments)} 条")


def git_commit_push():
    """git add + commit + push"""
    repo_dir = "/workspace"
    subprocess.run(["git", "add", "-A"], cwd=repo_dir, check=True)
    result = subprocess.run(["git", "status", "--porcelain"], cwd=repo_dir, capture_output=True, text=True)
    if not result.stdout.strip():
        print("git: 无变更")
        return False
    subprocess.run(
        ["git", "commit", "-m", f"feat: 更新购买记录 {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        cwd=repo_dir, check=True
    )
    push = subprocess.run(["git", "push"], cwd=repo_dir, capture_output=True, text=True)
    if push.returncode == 0:
        print("✅ git push 成功")
        return True
    else:
        print(f"❌ git push 失败: {push.stderr.strip()}")
        return False


# ==================== 已确认的购买记录 ====================
PURCHASE_RECORDS = [
    {
        "platform": "1688",
        "time": datetime(2026, 7, 12, 12, 11, 38),
        "product": "创意世界杯足球大力神杯",
        "spec": "500ml透明圆形足球杯(牛皮纸盒包装)",
        "qty": 1,
        "paid": 5.90,
        "order": "3311879379799327755",
        "shipment": "503-95444475-5097463",
        "ship_qty": 1,
        "stock": 0,
        "remark": "",
    },
    {
        "platform": "拼多多",
        "time": datetime(2026, 7, 18, 14, 11, 10),
        "product": "创意大力神杯世界杯足球玻璃啤酒杯",
        "spec": "500ML电镀圆球形大力神杯（牛皮纸盒包装）",
        "qty": 1,
        "paid": 20.00,
        "order": "260718-641864876461684",
        "shipment": "249-6387958-2579060",
        "ship_qty": 1,
        "stock": 0,
        "remark": "",
    },
    {
        "platform": "1688",
        "time": datetime(2026, 7, 21, 11, 0, 40),
        "product": "玻璃修复膏玻璃表面专业抛光",
        "spec": "50g+海绵",
        "qty": 2,
        "paid": 11.89,
        "order": "3313191614834136171",
        "shipment": "249-4866874-7926210",
        "ship_qty": 1,
        "stock": 1,
        "remark": "最低起订2个，卖出1个，库存1个",
    },
]

# ==================== 已确认的出货记录 ====================
SHIPMENT_RECORDS = [
    {
        "ship_platform": "Amazon日本站",
        "order_date": datetime(2026, 7, 11),
        "ship_date": datetime(2026, 7, 14),
        "product": "500ml啤酒杯耐热耐冷足球奖杯杯",
        "spec": "透明",
        "sku": "6E-4EHX-JF04",
        "qty": 1,
        "price": 2026,
        "tax": 184,
        "fee": 312,
        "revenue": 1714,
        "ship_order": "503-95444475-5097463",
        "purchase_ref": "3311879379799327755",
    },
    {
        "ship_platform": "Amazon",
        "order_date": datetime(2026, 7, 18),
        "ship_date": datetime(2026, 7, 21),
        "product": "500ml啤酒杯耐热耐冷足球奖杯杯",
        "spec": "琥珀色",
        "sku": "LK-RN2M-G3YX",
        "qty": 1,
        "price": 2026,
        "tax": 184,
        "fee": 312,
        "revenue": 1714,
        "ship_order": "249-6387958-2579060",
        "purchase_ref": "260718-641864876461684",
    },
    {
        "ship_platform": "Amazon日本站",
        "order_date": datetime(2026, 7, 20),
        "ship_date": datetime(2026, 7, 22),
        "product": "玻璃修复膏 50g 车用",
        "spec": "50g",
        "sku": "1X-GB0W-BY30",
        "qty": 1,
        "price": 1122,
        "tax": 102,
        "fee": 117,
        "revenue": 1005,
        "ship_order": "249-4866874-7926210",
        "purchase_ref": "3313191614834136171",
    },
]


if __name__ == "__main__":
    wb = load_workbook(XLSX)

    if "📦出货记录" not in wb.sheetnames:
        wb.create_sheet("📦出货记录")
        ws_ship = wb["📦出货记录"]
        ws_ship.merge_cells("A1:N1")
        ws_ship["A1"] = f"📦 出货记录表 ({YEAR}年)"
        ws_ship["A1"].font = FONT_TITLE
        ws_ship["A1"].fill = FILL_TITLE
        ws_ship["A1"].alignment = ALIGN_CENTER
        ws_ship.row_dimensions[1].height = 32
        for col_idx, h in enumerate(SHIPMENT_HEADERS, 1):
            cell = ws_ship.cell(row=2, column=col_idx, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = thin_border
        ws_ship.row_dimensions[2].height = 28
        wb.save(XLSX)

    ws = wb["🛒购买记录明细"]
    existing = get_existing_records(ws)
    existing_orders = {r["order"] for r in existing}

    new_purchases = []
    updated_purchases = []
    for r in PURCHASE_RECORDS:
        if r["order"] not in existing_orders:
            new_purchases.append(r)
        else:
            updated_purchases.append(r)

    ws_ship = wb["📦出货记录"]
    existing_shipments = get_existing_shipments(ws_ship)
    existing_ship_orders = {s["ship_order"] for s in existing_shipments}
    new_shipments = [s for s in SHIPMENT_RECORDS if s["ship_order"] not in existing_ship_orders]

    if new_purchases or new_shipments or updated_purchases:
        print(f"新增购买记录: {len(new_purchases)} 条")
        print(f"更新购买记录: {len(updated_purchases)} 条")
        print(f"新增出货记录: {len(new_shipments)} 条")
    else:
        print("无新增记录，仅重写表格")

    add_records(new_purchases + updated_purchases, new_shipments)

    git_commit_push()
