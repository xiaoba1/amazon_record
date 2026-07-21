"""
购买记录管理系统 - 新架构
四张表：进货表、销售表、库存表、关联视图
"""

import subprocess
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

YEAR = datetime.now().year
XLSX = f"归档表格/购买记录_{YEAR}.xlsx"

COLOR_HEADER = "4472C4"
COLOR_BORDER = "D9E1F2"
FONT_TITLE = Font(name="微软雅黑", bold=True, size=16, color="FFFFFF")
FONT_HEADER = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_TOTAL = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
FILL_TITLE = PatternFill(start_color="FF666666", end_color="FF666666", fill_type="solid")
FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FF666666", end_color="FF666666", fill_type="solid")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)


def auto_fit_columns(ws, headers, data_start_row=3, max_width=50, min_width=8):
    from openpyxl.utils import get_column_letter
    is_shipment = any("日元" in str(h) for h in headers)
    jpy_cols = (9, 10, 11, 12) if is_shipment else ()
    cny_cols = (7,) if not is_shipment else ()

    num_cols = len(headers)
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1] or ""))
        for row in range(data_start_row, min(502 + 1, data_start_row + 200)):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            if isinstance(val, (int, float)):
                if col_idx in jpy_cols:
                    s = f"JPY {int(val):,}"
                elif col_idx in cny_cols:
                    s = f"¥{val:,.2f}"
                else:
                    s = str(val)
            else:
                s = str(val)
            width = sum(2 if ord(c) > 127 else 1 for c in s)
            if width > max_len:
                max_len = width
        final_width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = final_width


PURCHASE_HEADERS = ["序号", "进货平台", "下单时间", "商品名称", "规格", "购买数量", "实付金额(元)", "订单号", "单价", "备注"]
SALES_HEADERS = ["序号", "销售平台", "订购日期", "发货日期", "商品名称", "规格", "SKU", "销售数量", "销售价(日元)", "税金(日元)", "手续费(日元)", "销售额(日元)", "订单号", "备注"]
INVENTORY_HEADERS = ["序号", "商品名称", "规格", "当前库存", "采购单价(元)", "最近采购时间", "最近销售时间", "备注"]
RELATION_HEADERS = ["序号", "进货订单号", "销售订单号", "商品名称", "规格", "采购数量", "销售数量", "库存", "采购金额(元)", "销售金额(日元)", "利润(日元)"]


def create_purchase_sheet(wb, records):
    if "📥进货表" in wb.sheetnames:
        idx = wb.sheetnames.index("📥进货表")
        old_ws = wb.worksheets[idx]
        wb.remove(old_ws)
    ws = wb.create_sheet("📥进货表")

    ws.merge_cells(f"A1:{chr(64 + len(PURCHASE_HEADERS))}1")
    title_cell = ws.cell(row=1, column=1, value=f"📥 进货表 ({YEAR}年)")
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    for col_idx, h in enumerate(PURCHASE_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    for i, rec in enumerate(records):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=rec["platform"])
        ws.cell(row=row, column=3, value=rec["time"])
        ws.cell(row=row, column=3).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row, column=4, value=rec["product"])
        ws.cell(row=row, column=5, value=rec["spec"])
        ws.cell(row=row, column=6, value=rec["qty"])
        ws.cell(row=row, column=7, value=rec["paid"])
        ws.cell(row=row, column=8, value=rec["order"])
        unit_price = round(rec["paid"] / rec["qty"], 2) if rec["qty"] > 0 else 0
        ws.cell(row=row, column=9, value=unit_price)
        if rec.get("remark"):
            ws.cell(row=row, column=10, value=rec["remark"])

        for col in range(1, len(PURCHASE_HEADERS) + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER if col in (1, 2, 3, 6, 7, 9) else ALIGN_LEFT
        ws.cell(row=row, column=7).number_format = "¥#,##0.00"
        ws.cell(row=row, column=9).number_format = "¥#,##0.00"
        ws.cell(row=row, column=6).number_format = "0"
        ws.row_dimensions[row].height = 22

    if records:
        total_row = 3 + len(records) + 1
        ws.merge_cells(f"A{total_row}:E{total_row}")
        ws.cell(row=total_row, column=1, value="📊 合计")
        ws.cell(row=total_row, column=1).font = FONT_TOTAL
        ws.cell(row=total_row, column=1).fill = FILL_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        total_paid = sum(float(r["paid"] or 0) for r in records)
        ws.cell(row=total_row, column=7, value=round(total_paid, 2))
        ws.cell(row=total_row, column=7).font = FONT_TOTAL
        ws.cell(row=total_row, column=7).fill = FILL_TOTAL
        ws.cell(row=total_row, column=7).number_format = "¥#,##0.00"
        ws.cell(row=total_row, column=7).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=7).border = thin_border
        ws.merge_cells(f"H{total_row}:J{total_row}")
        ws.cell(row=total_row, column=8, value=f"{len(records)} 笔进货")
        ws.cell(row=total_row, column=8).font = FONT_TOTAL
        ws.cell(row=total_row, column=8).fill = FILL_TOTAL
        ws.cell(row=total_row, column=8).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=8).border = thin_border
        ws.row_dimensions[total_row].height = 28

    auto_fit_columns(ws, PURCHASE_HEADERS)
    ws.freeze_panes = "A3"


def create_sales_sheet(wb, records):
    if "📤销售表" in wb.sheetnames:
        idx = wb.sheetnames.index("📤销售表")
        old_ws = wb.worksheets[idx]
        wb.remove(old_ws)
    ws = wb.create_sheet("📤销售表")

    last_col = chr(64 + len(SALES_HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value=f"📤 销售表 ({YEAR}年)")
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    for col_idx, h in enumerate(SALES_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    for i, rec in enumerate(records):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=rec["platform"])
        if rec["order_date"]:
            ws.cell(row=row, column=3, value=rec["order_date"])
            ws.cell(row=row, column=3).number_format = "yyyy-mm-dd"
        if rec["ship_date"]:
            ws.cell(row=row, column=4, value=rec["ship_date"])
            ws.cell(row=row, column=4).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=5, value=rec["product"])
        ws.cell(row=row, column=6, value=rec["spec"])
        ws.cell(row=row, column=7, value=rec["sku"])
        ws.cell(row=row, column=8, value=rec["qty"])
        ws.cell(row=row, column=9, value=rec["price"])
        ws.cell(row=row, column=10, value=rec["tax"])
        ws.cell(row=row, column=11, value=rec["fee"])
        ws.cell(row=row, column=12, value=rec["revenue"])
        ws.cell(row=row, column=13, value=rec["order"])
        if rec.get("remark"):
            ws.cell(row=row, column=14, value=rec["remark"])

        for col in range(1, len(SALES_HEADERS) + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER if col in (1, 2, 3, 4, 7, 8, 9, 10, 11, 12, 13) else ALIGN_LEFT
        ws.cell(row=row, column=9).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=10).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=11).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=12).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=8).number_format = "0"
        ws.row_dimensions[row].height = 22

    if records:
        total_row = 3 + len(records) + 1
        ws.merge_cells(f"A{total_row}:D{total_row}")
        ws.cell(row=total_row, column=1, value="📊 合计")
        ws.cell(row=total_row, column=1).font = FONT_TOTAL
        ws.cell(row=total_row, column=1).fill = FILL_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        total_revenue = sum(float(r["revenue"] or 0) for r in records)
        ws.merge_cells(f"I{total_row}:L{total_row}")
        ws.cell(row=total_row, column=9, value=f"总销售额: JPY {int(total_revenue):,}")
        ws.cell(row=total_row, column=9).font = FONT_TOTAL
        ws.cell(row=total_row, column=9).fill = FILL_TOTAL
        ws.cell(row=total_row, column=9).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=9).border = thin_border
        ws.merge_cells(f"M{total_row}:N{total_row}")
        ws.cell(row=total_row, column=13, value=f"{len(records)} 笔销售")
        ws.cell(row=total_row, column=13).font = FONT_TOTAL
        ws.cell(row=total_row, column=13).fill = FILL_TOTAL
        ws.cell(row=total_row, column=13).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=13).border = thin_border
        ws.row_dimensions[total_row].height = 28

    auto_fit_columns(ws, SALES_HEADERS)
    ws.freeze_panes = "A3"


def create_inventory_sheet(wb, purchases, sales):
    if "📦库存表" in wb.sheetnames:
        idx = wb.sheetnames.index("📦库存表")
        old_ws = wb.worksheets[idx]
        wb.remove(old_ws)
    ws = wb.create_sheet("📦库存表")

    ws.merge_cells(f"A1:{chr(64 + len(INVENTORY_HEADERS))}1")
    title_cell = ws.cell(row=1, column=1, value=f"📦 库存表 ({YEAR}年)")
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    for col_idx, h in enumerate(INVENTORY_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    from difflib import SequenceMatcher

    inventory = {}
    for p in purchases:
        key = (p["product"], p["spec"])
        if key not in inventory:
            inventory[key] = {
                "product": p["product"],
                "spec": p["spec"],
                "qty": 0,
                "paid": 0,
                "last_purchase": p["time"],
                "last_sale": None,
            }
        inventory[key]["qty"] += p["qty"]
        inventory[key]["paid"] += p["paid"]
        if p["time"] > inventory[key]["last_purchase"]:
            inventory[key]["last_purchase"] = p["time"]

    for s in sales:
        matched_key = None
        best_score = 0
        for p_key in inventory:
            p_product, p_spec = p_key
            score1 = SequenceMatcher(None, s["product"], p_product).ratio()
            score2 = SequenceMatcher(None, s["spec"], p_spec).ratio() if s["spec"] and p_spec else 0
            score = score1 * 0.7 + score2 * 0.3
            if score > best_score and score > 0.3:
                best_score = score
                matched_key = p_key

        if matched_key:
            inventory[matched_key]["qty"] -= s["qty"]
            if s["order_date"] > (inventory[matched_key]["last_sale"] or datetime.min):
                inventory[matched_key]["last_sale"] = s["order_date"]

    items = sorted(inventory.values(), key=lambda x: x["qty"], reverse=True)
    for i, item in enumerate(items):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=item["product"])
        ws.cell(row=row, column=3, value=item["spec"])
        ws.cell(row=row, column=4, value=item["qty"])

        total_purchases = sum(p["qty"] for p in purchases if (p["product"], p["spec"]) == (item["product"], item["spec"]))
        avg_price = round(item["paid"] / total_purchases, 2) if total_purchases > 0 else 0
        ws.cell(row=row, column=5, value=avg_price)

        if item["last_purchase"]:
            ws.cell(row=row, column=6, value=item["last_purchase"])
            ws.cell(row=row, column=6).number_format = "yyyy-mm-dd"
        if item["last_sale"]:
            ws.cell(row=row, column=7, value=item["last_sale"])
            ws.cell(row=row, column=7).number_format = "yyyy-mm-dd"

        stock_status = ""
        if item["qty"] > 0:
            stock_status = f"库存充足 ({item['qty']}个)"
        else:
            stock_status = "已售罄"
        ws.cell(row=row, column=8, value=stock_status)

        for col in range(1, len(INVENTORY_HEADERS) + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER if col in (1, 4, 5, 6, 7) else ALIGN_LEFT
        ws.cell(row=row, column=5).number_format = "¥#,##0.00"
        ws.cell(row=row, column=4).number_format = "0"
        ws.row_dimensions[row].height = 22

    if items:
        total_row = 3 + len(items) + 1
        ws.merge_cells(f"A{total_row}:C{total_row}")
        ws.cell(row=total_row, column=1, value="📊 库存合计")
        ws.cell(row=total_row, column=1).font = FONT_TOTAL
        ws.cell(row=total_row, column=1).fill = FILL_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        total_stock = sum(item["qty"] for item in items)
        ws.cell(row=total_row, column=4, value=total_stock)
        ws.cell(row=total_row, column=4).font = FONT_TOTAL
        ws.cell(row=total_row, column=4).fill = FILL_TOTAL
        ws.cell(row=total_row, column=4).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=4).border = thin_border
        ws.merge_cells(f"E{total_row}:H{total_row}")
        ws.cell(row=total_row, column=5, value=f"{len(items)} 种商品")
        ws.cell(row=total_row, column=5).font = FONT_TOTAL
        ws.cell(row=total_row, column=5).fill = FILL_TOTAL
        ws.cell(row=total_row, column=5).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=5).border = thin_border
        ws.row_dimensions[total_row].height = 28

    auto_fit_columns(ws, INVENTORY_HEADERS)
    ws.freeze_panes = "A3"


def create_relation_sheet(wb, purchases, sales):
    if "🔗关联视图" in wb.sheetnames:
        idx = wb.sheetnames.index("🔗关联视图")
        old_ws = wb.worksheets[idx]
        wb.remove(old_ws)
    ws = wb.create_sheet("🔗关联视图")

    ws.merge_cells(f"A1:{chr(64 + len(RELATION_HEADERS))}1")
    title_cell = ws.cell(row=1, column=1, value=f"🔗 关联视图 ({YEAR}年)")
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    for col_idx, h in enumerate(RELATION_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    from difflib import SequenceMatcher

    relations = []
    product_groups = {}

    for p in purchases:
        key = p["product"]
        if key not in product_groups:
            product_groups[key] = {"purchases": [], "sales": []}
        product_groups[key]["purchases"].append(p)

    for s in sales:
        best_group = None
        best_score = 0
        for group_key in product_groups:
            score = SequenceMatcher(None, s["product"], group_key).ratio()
            if score > best_score and score > 0.3:
                best_score = score
                best_group = group_key

        if best_group:
            product_groups[best_group]["sales"].append(s)
        else:
            product_groups[s["product"]] = {"purchases": [], "sales": [s]}

    for group_key, data in product_groups.items():
        if not data["purchases"] or not data["sales"]:
            continue

        pur_qty = sum(p["qty"] for p in data["purchases"])
        sale_qty = sum(s["qty"] for s in data["sales"])
        stock = pur_qty - sale_qty
        pur_amount = sum(p["paid"] for p in data["purchases"])
        sale_amount = sum(s["revenue"] for s in data["sales"])
        profit = sale_amount - (pur_amount * 20)

        for s in data["sales"]:
            relations.append({
                "purchase_order": ", ".join(p["order"] for p in data["purchases"]),
                "sale_order": s["order"],
                "product": group_key,
                "spec": data["purchases"][0]["spec"],
                "pur_qty": pur_qty,
                "sale_qty": sale_qty,
                "stock": stock,
                "pur_amount": pur_amount,
                "sale_amount": sale_amount,
                "profit": profit,
            })

    for i, rel in enumerate(relations):
        row = 3 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=rel["purchase_order"])
        ws.cell(row=row, column=3, value=rel["sale_order"])
        ws.cell(row=row, column=4, value=rel["product"])
        ws.cell(row=row, column=5, value=rel["spec"])
        ws.cell(row=row, column=6, value=rel["pur_qty"])
        ws.cell(row=row, column=7, value=rel["sale_qty"])
        ws.cell(row=row, column=8, value=rel["stock"])
        ws.cell(row=row, column=9, value=rel["pur_amount"])
        ws.cell(row=row, column=10, value=rel["sale_amount"])
        ws.cell(row=row, column=11, value=rel["profit"])

        for col in range(1, len(RELATION_HEADERS) + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER if col in (1, 6, 7, 8, 9, 10, 11) else ALIGN_LEFT
        ws.cell(row=row, column=9).number_format = "¥#,##0.00"
        ws.cell(row=row, column=10).number_format = '"JPY" #,##0'
        ws.cell(row=row, column=11).number_format = '"JPY" #,##0'
        ws.row_dimensions[row].height = 22

    if relations:
        total_row = 3 + len(relations) + 1
        ws.merge_cells(f"A{total_row}:E{total_row}")
        ws.cell(row=total_row, column=1, value="📊 合计")
        ws.cell(row=total_row, column=1).font = FONT_TOTAL
        ws.cell(row=total_row, column=1).fill = FILL_TOTAL
        ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER
        total_pur = sum(r["pur_amount"] for r in relations)
        total_sale = sum(r["sale_amount"] for r in relations)
        total_profit = sum(r["profit"] for r in relations)
        ws.cell(row=total_row, column=9, value=round(total_pur, 2))
        ws.cell(row=total_row, column=9).font = FONT_TOTAL
        ws.cell(row=total_row, column=9).fill = FILL_TOTAL
        ws.cell(row=total_row, column=9).number_format = "¥#,##0.00"
        ws.cell(row=total_row, column=9).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=9).border = thin_border
        ws.cell(row=total_row, column=10, value=int(total_sale))
        ws.cell(row=total_row, column=10).font = FONT_TOTAL
        ws.cell(row=total_row, column=10).fill = FILL_TOTAL
        ws.cell(row=total_row, column=10).number_format = '"JPY" #,##0'
        ws.cell(row=total_row, column=10).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=10).border = thin_border
        ws.cell(row=total_row, column=11, value=int(total_profit))
        ws.cell(row=total_row, column=11).font = FONT_TOTAL
        ws.cell(row=total_row, column=11).fill = FILL_TOTAL
        ws.cell(row=total_row, column=11).number_format = '"JPY" #,##0'
        ws.cell(row=total_row, column=11).alignment = ALIGN_CENTER
        ws.cell(row=total_row, column=11).border = thin_border
        ws.row_dimensions[total_row].height = 28

    auto_fit_columns(ws, RELATION_HEADERS)
    ws.freeze_panes = "A3"


PURCHASE_DATA = [
    {
        "platform": "1688",
        "time": datetime(2026, 7, 12, 12, 11, 38),
        "product": "创意世界杯足球大力神杯",
        "spec": "500ml透明圆形足球杯(牛皮纸盒包装)",
        "qty": 1,
        "paid": 5.90,
        "order": "3311879379799327755",
        "remark": "",
    },
    {
        "platform": "拼多多",
        "time": datetime(2026, 7, 18, 14, 11, 10),
        "product": "创意大力神杯世界杯足球玻璃啤酒杯",
        "spec": "500ML电镀圆球形大力神杯（牛皮纸盒包装）",
        "qty": 1,
        "paid": 20.00,
        "order": "260718-641864876461684",
        "remark": "",
    },
    {
        "platform": "1688",
        "time": datetime(2026, 7, 21, 11, 0, 40),
        "product": "玻璃修复膏玻璃表面专业抛光",
        "spec": "50g+海绵",
        "qty": 2,
        "paid": 11.89,
        "order": "3313191614834136171",
        "remark": "最低起订2个",
    },
]

SALES_DATA = [
    {
        "platform": "Amazon日本站",
        "order_date": datetime(2026, 7, 11),
        "ship_date": datetime(2026, 7, 14),
        "product": "500ml啤酒杯耐热耐冷足球奖杯杯",
        "spec": "透明",
        "sku": "6E-4EHX-JF04",
        "qty": 1,
        "price": 2026,
        "tax": 184,
        "fee": 312,
        "revenue": 1714,
        "order": "503-95444475-5097463",
        "remark": "",
    },
    {
        "platform": "Amazon",
        "order_date": datetime(2026, 7, 18),
        "ship_date": datetime(2026, 7, 21),
        "product": "500ml啤酒杯耐热耐冷足球奖杯杯",
        "spec": "琥珀色",
        "sku": "LK-RN2M-G3YX",
        "qty": 1,
        "price": 2026,
        "tax": 184,
        "fee": 312,
        "revenue": 1714,
        "order": "249-6387958-2579060",
        "remark": "",
    },
    {
        "platform": "Amazon日本站",
        "order_date": datetime(2026, 7, 20),
        "ship_date": datetime(2026, 7, 22),
        "product": "玻璃修复膏 50g 车用",
        "spec": "50g",
        "sku": "1X-GB0W-BY30",
        "qty": 1,
        "price": 1122,
        "tax": 102,
        "fee": 117,
        "revenue": 1005,
        "order": "249-4866874-7926210",
        "remark": "",
    },
]


def main():
    import os
    os.makedirs("归档表格", exist_ok=True)

    wb = load_workbook(XLSX) if os.path.exists(XLSX) else None
    if wb is None:
        from openpyxl import Workbook
        wb = Workbook()

    old_sheets = ["🛒购买记录明细", "📦出货记录", "📊月度统计", "📈年度统计", "🏪平台分布"]
    for name in old_sheets:
        if name in wb.sheetnames:
            idx = wb.sheetnames.index(name)
            std = wb.worksheets[idx]
            wb.remove(std)

    create_purchase_sheet(wb, PURCHASE_DATA)
    create_sales_sheet(wb, SALES_DATA)
    create_inventory_sheet(wb, PURCHASE_DATA, SALES_DATA)
    create_relation_sheet(wb, PURCHASE_DATA, SALES_DATA)

    if "Sheet" in wb.sheetnames:
        idx = wb.sheetnames.index("Sheet")
        std = wb.worksheets[idx]
        wb.remove(std)

    wb.save(XLSX)
    print(f"✅ 新架构表格已生成: {XLSX}")
    print(f"   📥进货表: {len(PURCHASE_DATA)} 条")
    print(f"   📤销售表: {len(SALES_DATA)} 条")

    git_commit_push()


def git_commit_push():
    repo_dir = "/workspace"
    subprocess.run(["git", "add", "-A"], cwd=repo_dir, check=True)
    subprocess.run(["git", "commit", "-m", f"feat: 重构为新架构 进货表/销售表/库存表/关联视图 {datetime.now().strftime('%Y-%m-%d %H:%M')}"], cwd=repo_dir, check=True)
    subprocess.run(["git", "push"], cwd=repo_dir, check=True)
    print("✅ git push 成功")


if __name__ == "__main__":
    main()
